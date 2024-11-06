import time
import datetime
from pathlib import Path
from time import perf_counter

from selenium.common import TimeoutException
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC


# 讀取xlsx檔案
import pandas as pd

from components.get_webdriver import driver
from components.table_data_cleanse import iselement
from components.utils import wait_for

from openpyxl import load_workbook
from tenacity import retry, stop_after_attempt
import copy

# 加载Excel文件
# workbook = load_workbook('../data.xlsx')
#
# # 选择工作表
# sheet = workbook.active  # 或者使用 workbook['工作表名']
# a = 0
# datas = ['https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7495344726565423350',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7494227701041957519',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7494008919956489925',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7495155741722184578',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7494001242412779805',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7495519059014224725',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7494019419359643757',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7493991856232433905',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7494246429697280709',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7495190556016544190',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7495279983602338736',
#          'https://affiliate.tiktokglobalshop.com/connection/creator/detail?cid=7495208385472203441']

# 读取数据
# for row in sheet.iter_rows(values_only=True):
#     if a != 0:
#         datas.append(row[-1])
#     a += 1
# print(datas)
# driver.get('https://affiliate.tiktokglobalshop.com/connection/creator?enter_from=affiliate_find_creators&shop_region=US')
# wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'table tbody tr')))
# driver.find_element(By.CSS_SELECTOR, 'table tbody tr').click()
# b = 0
# flag = True

# c = copy.deepcopy(datas)
# @retry(stop=stop_after_attempt(len(datas) * 2))
# def a(_datas):
#     global b
#     for i in _datas[b:]:
#         ee = int(input('輸入數字：'))
#         print('retry')
#         ert = 1 / ee
#         c[b] = b
#         b += 1
#     return c
#
#
# print(a(datas))
# dict_data = [
#     {'name': 'test1\ntest2'},
#     {'name': 'test2'},
#     {'name': 'test3'}
# ]
#
# df = pd.DataFrame(dict_data)
# df.to_excel('test.xlsx', index=False)

# def get_data():
#     dict_data[0] = dict_data[0] | {'age': 18}
# get_data()
# print(dict_data)
# count = 0

# def get_data(_datas):
#     copy_datas = copy.deepcopy(_datas)
#     for i in _datas:
#         detail_url = i
#         # 打開一個新標簽頁面
#         driver.get(detail_url)
#         # 等待頁面載入完成
#         try:
#             wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, '.pcm-pc-content .pcm-pc-legend')))
#             deal_amount = driver.find_element(By.CSS_SELECTOR, '.pcm-pc-content .pcm-pc-legend').find_elements(
#                 By.CSS_SELECTOR, '.ecom-data-overflow-text-container')
#             deal_list = [i.text for i in deal_amount]
#             _list = driver.find_elements(By.CSS_SELECTOR, '[data-e2e="0bc7b49d-b8b3-02d5"]')
#             dict_data = {
#                 # 獲取詳細頁 個人簡介
#                 'intro': iselement(driver, By.CSS_SELECTOR, '[data-e2e="2e9732e6-4d06-458d"]'),
#                 # 獲取詳細頁 預發佈率
#                 'kbps': _list[4].text,
#                 # 獲取粉絲數,
#                 'follower_count': iselement(driver, By.CSS_SELECTOR, '[data-e2e="7aed0dd7-48ba-6932"]'),
#                 # 獲取佣金率
#                 'commission': _list[5].text,
#                 # 獲取每个销售渠道的商品交易总额
#                 'deal_amount': deal_list
#             }
#             global count
#             _datas = _datas[count + 1:]
#             copy_datas[count] = dict_data
#             count += 1
#         except TimeoutException as e:
#             print('頁面載入超時')
#             print('無法取得資料,出現滑塊驗證')
#     return copy_datas
#
#
# print(get_data(datas))
# print(datas)

# df = pd.read_excel('../data.xlsx')
# df = df.drop_duplicates(subset=['name'])
# df.to_excel('../data1.xlsx', index=False)
# print(len(df))
# 当前日期
# today = datetime.date.today()
# # 创建一个 timedelta 对象，表示 10 天
# ten_days = datetime.timedelta(days=10)
# # 加上 10 天
# future_date = today + ten_days
# print(future_date.strftime('%m/%d/%Y') + '1111')  # 输出: 当前日期加 10 天

# 數據切片
a = [
    {'name':'xiao1'},
    {'age':'12'}
]

# b = [
#     {'name':'xiao2', 'age':'99', 'city':'beijing'},
#     {'name': 'xiao2', 'age': '88', 'city': 'beijing'},
# ]
#
# da = pd.DataFrame(a)
# db = pd.DataFrame(b)
#
# # 合并数据
# df = pd.concat([db, da], ignore_index=True)
# # df.drop_duplicates(subset=['name'], inplace=True)
# print(df)


driver.get('https://affiliate.tiktokglobalshop.com/connection/target-invitation/create?creator_ids[0]=7493991992885676258&creator_ids[1]=7493996432515500778&creator_ids[2]=7493997719687367005&creator_ids[3]=7494000761953618894&creator_ids[4]=7494004143427192413&creator_ids[5]=7494004464331556606&creator_ids[6]=7494008989740795926&creator_ids[7]=7494009124373104689&creator_ids[8]=7494009403495909109&creator_ids[9]=7494010320652240592&creator_ids[10]=7494010344646673522&creator_ids[11]=7494010365195289615&creator_ids[12]=7494010854056690785&creator_ids[13]=7494012999735675318&creator_ids[14]=7494013745611311226&creator_ids[15]=7494016216561387202&creator_ids[16]=7494018542024035197&creator_ids[17]=7494018584035887164&creator_ids[18]=7494021806615528839&creator_ids[19]=7494022088330282017&enter_from=affiliate_crm&shop_region=US')
time.sleep(5)
driver.execute_script("Array.from(document.querySelectorAll('[role=region]')).forEach(item => item.style.display = 'block')")
# driver.execute_script("document.querySelector('#target_complete_details_message_input').value = 'adawdawd'")
driver.find_element(By.ID, 'target_complete_details_message_input').send_keys('我是输入的内容')

input()